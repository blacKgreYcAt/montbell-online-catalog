#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def create_public_template():
    """Create public version product template"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Products"

    # Headers
    headers = [
        "modelNumber", "name", "nameEn", "category",
        "description", "features", "specifications",
        "sizes", "weight", "price", "colors", "season", "badge", "sampleColor", "pageNumber"
    ]

    ws.append(headers)

    # Header style
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Example data
    example_data = [
        [
            "1128850",
            "DRY-TEC FULL-ZIP RAIN PANTS MEN'S",
            "DRY-TEC FULL-ZIP RAIN PANTS MEN'S",
            "ALLWEATHER",
            "Weight: 280g, Material: 50D tear-resistant nylon",
            "Waterproof breathable, lightweight design",
            "50D tear-resistant nylon, DRYTEC 3L",
            "S,M,L,XL",
            "280g",
            "4040",
            "BK,WH,RD",
            "SS",
            "",
            "BK",
            "1"
        ],
        [
            "1101770",
            "PERMAFROST Down Parka Men",
            "PERMAFROST DOWN PARKA MEN'S",
            "INSULATION",
            "Down parka, waterproof breathable shell",
            "Waterproof, box construction, detachable hood",
            "2-layer DRYTEC, 800 fill power down",
            "S,M,L,XL,XXL",
            "550g",
            "11630",
            "BK,DTL,MST",
            "FW",
            "",
            "BK",
            "2"
        ]
    ]

    for row in example_data:
        ws.append(row)

    # Column widths
    column_widths = {
        "A": 12, "B": 25, "C": 30, "D": 15, "E": 35,
        "F": 30, "G": 35, "H": 20, "I": 10, "J": 10,
        "K": 15, "L": 5, "M": 10, "N": 10, "O": 5
    }

    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # Data row style
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    # Info sheet
    ws_info = wb.create_sheet("Information")
    info_content = [
        ["Field Description - Public Version"],
        [],
        ["modelNumber", "Product model (Required)", "e.g. 1128850"],
        ["name", "Chinese name (Required)", "e.g. DRY-TEC FULL-ZIP RAIN PANTS"],
        ["nameEn", "English name (Required)", "e.g. DRY-TEC FULL-ZIP RAIN PANTS"],
        ["category", "Category (Required)", "e.g. ALLWEATHER, INSULATION"],
        ["description", "Chinese description", "Detailed product description"],
        ["features", "Chinese features", "Main features, use \\n for line break"],
        ["specifications", "Chinese specifications", "Technical specs, use \\n for line break"],
        ["sizes", "Sizes", "Multiple separated by comma, e.g. S,M,L,XL"],
        ["weight", "Weight", "e.g. 280g"],
        ["price", "Price (Public)", "Numbers only, e.g. 4040 (means NT$4040)"],
        ["colors", "Color codes", "Multiple separated by comma. Standard codes: BK,WH,RD,BL,GR,YE,PK,BR,GY,BE,NV,DGY,RBL,GN,OR,TN,KH"],
        ["season", "Season", "SS=Spring/Summer, FW=Fall/Winter"],
        ["badge", "Badge (Optional)", "e.g. Special, 26FW, 27SS (leave empty if none)"],
        ["sampleColor", "Sample color", "e.g. BK (one color code)"],
        ["pageNumber", "Page number", "Catalog page number"],
    ]

    for row in info_content:
        ws_info.append(row)

    ws_info.column_dimensions["A"].width = 15
    ws_info.column_dimensions["B"].width = 30
    ws_info.column_dimensions["C"].width = 40

    wb.save("products_template.xlsx")
    print("Created: products_template.xlsx (Public Version)")

def create_internal_template():
    """Create internal version product template"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Products"

    # Headers
    headers = [
        "modelNumber", "name", "nameEn", "category",
        "description", "features", "specifications",
        "sizes", "weight", "price", "colors", "season", "sampleColor"
    ]

    ws.append(headers)

    # Header style
    header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Example data
    example_data = [
        [
            "1101770",
            "PERMAFROST Down Parka Men",
            "PERMAFROST DOWN PARKA MEN'S",
            "INSULATION",
            "Down parka, waterproof breathable shell, 800 fill power EX down",
            "Waterproof, breathable DRYTEC, box construction, detachable hood",
            "2-layer DRYTEC (high breathability)\n30-denier nylon ripstop\nLining: 20-denier Ballistic nylon\nInsulation: 800 fill power EX down",
            "S,M,L,XL,XXL",
            "550g",
            "NT$11,630",
            "BK,DTL,MST",
            "FW",
            "BK"
        ],
        [
            "1101771",
            "PERMAFROST Down Parka Women",
            "PERMAFROST DOWN PARKA WOMEN'S",
            "INSULATION",
            "Down parka, waterproof breathable shell, 800 fill power EX down",
            "Waterproof, breathable DRYTEC, box construction",
            "2-layer DRYTEC (high breathability)\n30-denier nylon ripstop",
            "XS,S,M,L,XL",
            "500g",
            "NT$11,190",
            "BK,DTL",
            "FW",
            "BK"
        ]
    ]

    for row in example_data:
        ws.append(row)

    # Column widths
    column_widths = {
        "A": 12, "B": 25, "C": 30, "D": 15, "E": 35,
        "F": 30, "G": 35, "H": 20, "I": 10, "J": 12,
        "K": 15, "L": 5, "M": 10
    }

    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # Data row style
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    # Info sheet
    ws_info = wb.create_sheet("Information")
    info_content = [
        ["Field Description - Internal Version"],
        [],
        ["modelNumber", "Product model (Required)", "e.g. 1101770"],
        ["name", "Chinese name (Required)", "e.g. PERMAFROST Down Parka Men"],
        ["nameEn", "English name (Required)", "e.g. PERMAFROST DOWN PARKA MEN'S"],
        ["category", "Category (Required)", "e.g. INSULATION"],
        ["description", "Chinese description", "Detailed product description"],
        ["features", "Chinese features", "Main features, use \\n for line break"],
        ["specifications", "Chinese specifications", "Technical specs, use \\n for line break"],
        ["sizes", "Sizes", "Multiple separated by comma, e.g. S,M,L,XL"],
        ["weight", "Weight", "e.g. 550g"],
        ["price", "Price (Internal - Taiwan Dollar)", "MUST include NT$ symbol, e.g. NT$11,630"],
        ["colors", "Color codes", "Multiple separated by comma. Standard codes: BK,WH,RD,BL,GR,YE,PK,BR,GY,BE,NV,DGY,RBL,GN,OR,TN,KH"],
        ["season", "Season", "Usually FW (Fall/Winter)"],
        ["sampleColor", "Sample color", "e.g. BK (one color code)"],
    ]

    for row in info_content:
        ws_info.append(row)

    ws_info.column_dimensions["A"].width = 15
    ws_info.column_dimensions["B"].width = 30
    ws_info.column_dimensions["C"].width = 40

    wb.save("products_internal_template.xlsx")
    print("Created: products_internal_template.xlsx (Internal Version)")

if __name__ == "__main__":
    create_public_template()
    create_internal_template()
    print("\nBoth templates created successfully!")
