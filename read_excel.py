#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel 商品資料轉換腳本
將 Excel 文件轉換為 products.json，自動檢測季節

使用方式：
  python read_excel.py [Excel文件路徑]

範例：
  python read_excel.py "27SS 正式_ 工作本連結資料27SS 資料檔.xlsx"
  python read_excel.py "27FW 正式_ 工作本連結資料27FW 資料檔.xlsx"

季節偵測：
  - 文件名包含 'SS' 或 'ss' → SS (春夏)
  - 文件名包含 'FW' 或 'fw' → FW (秋冬)
  - 預設值：SS
"""

import warnings
import sys
import os
import json
import re
from pathlib import Path

warnings.filterwarnings('ignore')

# Set stdout encoding to UTF-8
if sys.stdout.encoding != 'utf-8':
    import io
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

from openpyxl import load_workbook


def detect_season(file_path: str) -> str:
    """根據文件名偵測季節"""
    filename = os.path.basename(file_path).upper()
    if 'FW' in filename:
        return 'FW'
    elif 'SS' in filename:
        return 'SS'
    return 'SS'  # 預設值


def parse_colors(color_str: str) -> list:
    """解析顏色字符串"""
    if not color_str:
        return []
    # 分割多個顏色（用逗號或空白分隔）
    colors = [c.strip().upper() for c in str(color_str).replace('，', ',').split(',') if c.strip()]
    # 也處理空白分隔的情況
    if len(colors) == 1:
        parts = colors[0].split()
        if len(parts) > 1:
            colors = parts
    return colors


def parse_sizes(size_str: str) -> list:
    """解析尺寸字符串"""
    if not size_str:
        return []
    # 分割多個尺寸（用逗號或空白分隔）
    sizes = [s.strip() for s in str(size_str).replace('，', ',').split(',') if s.strip()]
    # 也處理空白分隔的情況
    if len(sizes) == 1:
        parts = sizes[0].split()
        if len(parts) > 1:
            sizes = parts
    return sizes


def convert_excel_to_json(file_path: str, output_path: str = 'public/products.json'):
    """轉換 Excel 為 products.json"""

    print(f"開始轉換: {file_path}")
    print(f"輸出路徑: {output_path}")

    season = detect_season(file_path)
    print(f"偵測季節: {season}\n")

    products = []

    try:
        wb = load_workbook(file_path, data_only=True)
        sheet_names = wb.sheetnames
        print(f"找到 {len(sheet_names)} 個工作表: {sheet_names}\n")

        for sheet_name in sheet_names:
            ws = wb[sheet_name]
            print(f"{'='*120}")
            print(f"工作表: {sheet_name}")
            print(f"最大行: {ws.max_row}, 最大列: {ws.max_column}\n")

            # 取得標題行
            headers = []
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=1, column=col_idx)
                headers.append(cell.value)

            print(f"欄位列表:")
            for i, h in enumerate(headers, 1):
                print(f"  {i}. {h}")
            print()

            # Excel 欄位映射（根據實際結構）
            # Col 1: 改款/推薦 (Badge)
            # Col 2: 系列 (Category)
            # Col 3: 型號 (Model Number)
            # Col 4: 品名 (Product Name)
            # Col 5: 色號名稱 (Colors)
            # Col 6: 尺寸 (Sizes)
            # Col 7: 訂價 (Price)
            # Col 8: 特色 (Description)
            # Col 9: 頁碼 (Page Number)

            print(f"開始解析商品數據...\n")

            # 解析數據行
            for row_idx in range(2, ws.max_row + 1):
                try:
                    # 根據列索引獲取數據
                    badge = ws.cell(row=row_idx, column=1).value
                    category = ws.cell(row=row_idx, column=2).value
                    model_number = ws.cell(row=row_idx, column=3).value
                    name = ws.cell(row=row_idx, column=4).value
                    colors_str = ws.cell(row=row_idx, column=5).value
                    sizes_str = ws.cell(row=row_idx, column=6).value
                    price = ws.cell(row=row_idx, column=7).value
                    description = ws.cell(row=row_idx, column=8).value
                    page_number = ws.cell(row=row_idx, column=9).value

                    # 驗證必要字段
                    if not model_number or not name:
                        continue

                    # 清理數據
                    model_number = str(model_number).strip()
                    name = str(name).strip()
                    category = str(category).strip() if category else ''

                    # 構建商品對象
                    product = {
                        'id': f"prod-{model_number}",
                        'modelNumber': model_number,
                        'name': name,
                        'category': category,
                        'season': season,  # 新增季節字段
                        'colors': parse_colors(colors_str),
                        'sizes': parse_sizes(sizes_str),
                        'isNew': False,
                        'badge': str(badge).strip() if badge and str(badge).strip() not in ['None', ''] else None,
                        'pageNumber': int(page_number) if page_number else row_idx - 1,
                    }

                    # 可選字段
                    if description:
                        product['description'] = str(description).strip()

                    if price:
                        try:
                            product['price'] = int(float(price))
                        except (ValueError, TypeError):
                            pass

                    products.append(product)
                    print(f"✓ Row {row_idx:4d}: {model_number:10s} - {name[:50]}")

                except Exception as e:
                    print(f"✗ Row {row_idx:4d}: 錯誤 - {str(e)}")
                    continue

        print(f"\n{'='*120}")
        print(f"轉換完成！")
        print(f"共轉換 {len(products)} 件商品\n")

        # 保存為 JSON
        os.makedirs(os.path.dirname(output_path) or '.', exist_ok=True)
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(products, f, ensure_ascii=False, indent=2)

        print(f"已保存到: {output_path}\n")

        # 顯示摘要
        categories = set(p['category'] for p in products if p['category'])
        colors_set = set()
        for p in products:
            colors_set.update(p.get('colors', []))

        print(f"摘要:")
        print(f"  季節: {season}")
        print(f"  總商品數: {len(products)}")
        print(f"  分類數: {len(categories)}")
        print(f"  顏色數: {len(colors_set)}")
        print(f"\n分類列表:")
        for cat in sorted(categories):
            count = len([p for p in products if p['category'] == cat])
            print(f"  {cat}: {count} 件")

        return products

    except Exception as e:
        import traceback
        print(f"錯誤: {e}")
        traceback.print_exc()
        return None


if __name__ == '__main__':
    # 從命令行參數獲取文件路徑
    if len(sys.argv) > 1:
        excel_file = sys.argv[1]
    else:
        # 預設文件
        excel_file = '27SS 正式_ 工作本連結資料27SS 資料檔.xlsx'

    if os.path.exists(excel_file):
        convert_excel_to_json(excel_file)
    else:
        print(f"錯誤：找不到文件 {excel_file}")
        print(f"使用方式: python read_excel.py [Excel文件路徑]")
